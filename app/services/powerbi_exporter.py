"""Power BI-friendly Excel exporter for InsightFlow AI.

This module stores validated analytical runs in a stable, append-only workbook.
Power BI Desktop can import the workbook and relate all sheets through run_id.

No LLM-generated number is trusted by this module. Callers must pass results
produced by the deterministic executor after quality-gate validation.
"""

from __future__ import annotations

import hashlib
import json
import os
import tempfile
import uuid
import zipfile
from xml.etree import ElementTree as ET
from dataclasses import asdict, dataclass, field, is_dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


WORKBOOK_VERSION = "1.0"
DEFAULT_OUTPUT_PATH = Path("powerbi_output") / "insightflow_powerbi.xlsx"

SHEET_HEADERS: dict[str, list[str]] = {
    "AnalysisRuns": [
        "run_id",
        "publication_key",
        "question",
        "response_language",
        "planner_source",
        "model_name",
        "intent",
        "metric",
        "dimension",
        "aggregation",
        "sort_direction",
        "result_limit",
        "visualization",
        "confidence",
        "source_rows",
        "filtered_rows",
        "validation_status",
        "quality_ready",
        "created_at_utc",
        "workbook_version",
    ],
    "AnalysisResults": [
        "run_id",
        "result_order",
        "label",
        "value",
        "result_rank",
        "metric",
        "dimension",
        "unit",
        "created_at_utc",
    ],
    "AIResponses": [
        "run_id",
        "language",
        "answer",
        "primary_finding",
        "calculation",
        "limitations",
        "follow_up_question",
        "reasoning_summary",
        "assumptions_json",
        "tool_steps_json",
        "created_at_utc",
    ],
    "QualityEvidence": [
        "run_id",
        "gate_name",
        "ready",
        "duplicate_rows",
        "missing_values",
        "usable_numeric_ratio",
        "validation_status",
        "warnings_json",
        "evidence_json",
        "created_at_utc",
    ],
    "SourceFiles": [
        "run_id",
        "filename",
        "file_type",
        "file_size_bytes",
        "page_count",
        "ocr_executed",
        "ocr_languages",
        "extraction_strategy",
        "extraction_confidence",
        "source_sha256",
        "created_at_utc",
    ],
}


@dataclass(frozen=True)
class PowerBIResultRow:
    """One chart-ready result row."""

    label: str
    value: float | int
    rank: int | None = None
    unit: str | None = None


@dataclass(frozen=True)
class PowerBIResponse:
    """Grounded final-response fields shown in Power BI."""

    language: str = "en"
    answer: str = ""
    primary_finding: str = ""
    calculation: str = ""
    limitations: str = ""
    follow_up_question: str = ""


@dataclass(frozen=True)
class PowerBIQualityEvidence:
    """Quality-gate evidence for one analytical run."""

    gate_name: str = "analysis"
    ready: bool = True
    duplicate_rows: int = 0
    missing_values: int = 0
    usable_numeric_ratio: float | None = None
    validation_status: str = "passed"
    warnings: Sequence[str] = field(default_factory=tuple)
    evidence: Mapping[str, Any] = field(default_factory=dict)


@dataclass(frozen=True)
class PowerBISourceFile:
    """Source-document provenance."""

    filename: str = ""
    file_type: str = ""
    file_size_bytes: int | None = None
    page_count: int | None = None
    ocr_executed: bool = False
    ocr_languages: str = ""
    extraction_strategy: str = ""
    extraction_confidence: float | None = None
    source_sha256: str = ""


@dataclass(frozen=True)
class PowerBIPublication:
    """Complete validated payload for one Power BI publication."""

    question: str
    planner_source: str
    model_name: str
    intent: str
    metric: str | None
    dimension: str | None
    aggregation: str
    visualization: str
    results: Sequence[PowerBIResultRow]
    response_language: str = "en"
    sort_direction: str | None = None
    result_limit: int | None = None
    confidence: float | None = None
    source_rows: int | None = None
    filtered_rows: int | None = None
    validation_status: str = "passed"
    quality_ready: bool = True
    ai_response: PowerBIResponse = field(default_factory=PowerBIResponse)
    quality: Sequence[PowerBIQualityEvidence] = field(default_factory=tuple)
    source_files: Sequence[PowerBISourceFile] = field(default_factory=tuple)
    reasoning_summary: str = ""
    assumptions: Sequence[str] = field(default_factory=tuple)
    tool_steps: Sequence[Mapping[str, Any]] = field(default_factory=tuple)
    publication_key: str | None = None
    run_id: str | None = None
    created_at_utc: datetime | None = None


@dataclass(frozen=True)
class PowerBIPublishResult:
    """Successful publication metadata returned to Streamlit."""

    run_id: str
    publication_key: str
    workbook_path: Path
    duplicate: bool
    result_rows_written: int


def _utc_now() -> datetime:
    return datetime.now(timezone.utc).replace(microsecond=0)


def _iso_utc(value: datetime | None) -> str:
    timestamp = value or _utc_now()
    if timestamp.tzinfo is None:
        timestamp = timestamp.replace(tzinfo=timezone.utc)
    return timestamp.astimezone(timezone.utc).replace(microsecond=0).isoformat()


def _json_default(value: Any) -> Any:
    if is_dataclass(value):
        return asdict(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Path):
        return str(value)
    if hasattr(value, "item"):
        return value.item()
    return str(value)


def _json_text(value: Any) -> str:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=_json_default,
    )


def _publication_key(publication: PowerBIPublication) -> str:
    if publication.publication_key:
        return publication.publication_key.strip()

    deterministic_payload = {
        "question": publication.question.strip(),
        "planner_source": publication.planner_source,
        "intent": publication.intent,
        "metric": publication.metric,
        "dimension": publication.dimension,
        "aggregation": publication.aggregation,
        "sort_direction": publication.sort_direction,
        "result_limit": publication.result_limit,
        "results": [asdict(row) for row in publication.results],
        "validation_status": publication.validation_status,
        "source_files": [asdict(item) for item in publication.source_files],
    }
    return hashlib.sha256(_json_text(deterministic_payload).encode("utf-8")).hexdigest()


def sha256_file(path: str | Path) -> str:
    """Return the SHA-256 checksum of a source file."""

    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _validate_publication(publication: PowerBIPublication) -> None:
    if not publication.question.strip():
        raise ValueError("question must not be empty")
    if not publication.planner_source.strip():
        raise ValueError("planner_source must not be empty")
    if not publication.intent.strip():
        raise ValueError("intent must not be empty")
    if not publication.aggregation.strip():
        raise ValueError("aggregation must not be empty")
    if not publication.results:
        raise ValueError("at least one validated result row is required")
    if not publication.quality_ready:
        raise ValueError("Power BI publication is blocked because quality_ready is false")
    if publication.validation_status.casefold() not in {"passed", "pass", "ready", "validated"}:
        raise ValueError("Power BI publication requires a passed validation_status")

    for index, row in enumerate(publication.results, start=1):
        if not row.label.strip():
            raise ValueError(f"result row {index} has an empty label")
        if isinstance(row.value, bool) or not isinstance(row.value, (int, float)):
            raise TypeError(f"result row {index} value must be numeric")
        if row.rank is not None and row.rank < 1:
            raise ValueError(f"result row {index} rank must be positive")


def _initialize_workbook(path: Path) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_name, headers in SHEET_HEADERS.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(name="Calibri", size=11, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 24

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _ensure_workbook_schema(workbook: Any) -> None:
    for sheet_name, expected_headers in SHEET_HEADERS.items():
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
            sheet.append(expected_headers)
        sheet = workbook[sheet_name]
        actual_headers = [sheet.cell(row=1, column=index).value for index in range(1, len(expected_headers) + 1)]
        if actual_headers != expected_headers:
            raise ValueError(
                f"Workbook sheet '{sheet_name}' has an incompatible schema. "
                f"Expected {expected_headers}, received {actual_headers}."
            )


def _existing_publications(workbook: Any) -> dict[str, str]:
    sheet = workbook["AnalysisRuns"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    key_column = headers["publication_key"]
    run_column = headers["run_id"]
    existing: dict[str, str] = {}
    for row_index in range(2, sheet.max_row + 1):
        key = sheet.cell(row=row_index, column=key_column).value
        run_id = sheet.cell(row=row_index, column=run_column).value
        if key:
            existing[str(key)] = str(run_id)
    return existing


def _append_rows(workbook: Any, rows_by_sheet: Mapping[str, Sequence[Sequence[Any]]]) -> None:
    for sheet_name, rows in rows_by_sheet.items():
        sheet = workbook[sheet_name]
        for row in rows:
            sheet.append(list(row))


def _safe_table_name(sheet_name: str) -> str:
    return f"tbl{sheet_name}"


def _style_workbook(workbook: Any) -> None:
    widths: dict[str, dict[str, float]] = {
        "AnalysisRuns": {
            "A": 38, "B": 66, "C": 48, "D": 18, "E": 18, "F": 18,
            "G": 16, "H": 18, "I": 18, "J": 16, "K": 18, "L": 14,
            "M": 16, "N": 14, "O": 14, "P": 14, "Q": 20, "R": 14,
            "S": 26, "T": 18,
        },
        "AnalysisResults": {
            "A": 38, "B": 14, "C": 32, "D": 18, "E": 14, "F": 18,
            "G": 18, "H": 14, "I": 26,
        },
        "AIResponses": {
            "A": 38, "B": 14, "C": 72, "D": 54, "E": 48, "F": 48,
            "G": 44, "H": 64, "I": 56, "J": 56, "K": 26,
        },
        "QualityEvidence": {
            "A": 38, "B": 24, "C": 12, "D": 18, "E": 18, "F": 22,
            "G": 20, "H": 56, "I": 70, "J": 26,
        },
        "SourceFiles": {
            "A": 38, "B": 44, "C": 18, "D": 18, "E": 14, "F": 16,
            "G": 20, "H": 30, "I": 24, "J": 66, "K": 26,
        },
    }

    for sheet_name, headers in SHEET_HEADERS.items():
        sheet = workbook[sheet_name]
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False
        sheet.auto_filter.ref = None
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(name="Calibri", size=11, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column, width in widths[sheet_name].items():
            sheet.column_dimensions[column].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = Font(name="Calibri", size=11)

        if sheet.max_row >= 2:
            table_ref = f"A1:{get_column_letter(len(headers))}{sheet.max_row}"
            table_name = _safe_table_name(sheet_name)
            if table_name in sheet.tables:
                del sheet.tables[table_name]
            table = Table(displayName=table_name, ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)

        header_index = {name: index + 1 for index, name in enumerate(headers)}
        for field_name in ("confidence", "usable_numeric_ratio", "extraction_confidence"):
            if field_name in header_index:
                column_index = header_index[field_name]
                for row_index in range(2, sheet.max_row + 1):
                    sheet.cell(row=row_index, column=column_index).number_format = "0.0%"
        if "value" in header_index:
            column_index = header_index["value"]
            for row_index in range(2, sheet.max_row + 1):
                sheet.cell(row=row_index, column=column_index).number_format = '#,##0.00;[Red](#,##0.00);-'


def _normalize_styles_xml_font_order(xlsx_path: Path) -> None:
    """Normalize SpreadsheetML font child order for strict Excel readers."""

    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ET.register_namespace("", namespace)
    order = {
        "name": 0, "charset": 1, "family": 2, "b": 3, "i": 4,
        "strike": 5, "outline": 6, "shadow": 7, "condense": 8,
        "extend": 9, "sz": 10, "color": 11, "u": 12,
        "vertAlign": 13, "scheme": 14,
    }

    with zipfile.ZipFile(xlsx_path, "r") as source:
        entries = {name: source.read(name) for name in source.namelist()}

    styles_name = "xl/styles.xml"
    root = ET.fromstring(entries[styles_name])
    fonts = root.find(f"{{{namespace}}}fonts")
    if fonts is not None:
        for font in fonts.findall(f"{{{namespace}}}font"):
            children = list(font)
            children.sort(
                key=lambda child: order.get(child.tag.rsplit("}", 1)[-1], 99)
            )
            font[:] = children

    entries[styles_name] = ET.tostring(
        root, encoding="utf-8", xml_declaration=False
    )

    normalized_path = xlsx_path.with_suffix(".normalized.xlsx")
    with zipfile.ZipFile(
        normalized_path, "w", compression=zipfile.ZIP_DEFLATED
    ) as target:
        for name, payload in entries.items():
            target.writestr(name, payload)

    os.replace(normalized_path, xlsx_path)


def _atomic_save(workbook: Any, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    file_descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{output_path.stem}_",
        suffix=".xlsx",
        dir=output_path.parent,
    )
    os.close(file_descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        _normalize_styles_xml_font_order(temporary_path)
        os.replace(temporary_path, output_path)
    finally:
        temporary_path.unlink(missing_ok=True)


def publish_to_powerbi_workbook(
    publication: PowerBIPublication,
    output_path: str | Path = DEFAULT_OUTPUT_PATH,
    *,
    allow_duplicate: bool = False,
) -> PowerBIPublishResult:
    """Append one validated analytical run to the Power BI workbook.

    Duplicate prevention is based on a deterministic publication_key. Passing
    allow_duplicate=True creates another run with the same analytical payload.
    """

    _validate_publication(publication)
    path = Path(output_path)
    if not path.exists():
        _initialize_workbook(path)

    workbook = load_workbook(path)
    _ensure_workbook_schema(workbook)

    publication_key = _publication_key(publication)
    existing = _existing_publications(workbook)
    if publication_key in existing and not allow_duplicate:
        return PowerBIPublishResult(
            run_id=existing[publication_key],
            publication_key=publication_key,
            workbook_path=path.resolve(),
            duplicate=True,
            result_rows_written=0,
        )

    run_id = publication.run_id or str(uuid.uuid4())
    created_at = _iso_utc(publication.created_at_utc)

    response = publication.ai_response
    quality_rows = publication.quality or (
        PowerBIQualityEvidence(
            gate_name="analysis",
            ready=publication.quality_ready,
            validation_status=publication.validation_status,
        ),
    )

    rows_by_sheet: dict[str, list[list[Any]]] = {
        "AnalysisRuns": [[
            run_id,
            publication_key,
            publication.question,
            publication.response_language,
            publication.planner_source,
            publication.model_name,
            publication.intent,
            publication.metric,
            publication.dimension,
            publication.aggregation,
            publication.sort_direction,
            publication.result_limit,
            publication.visualization,
            publication.confidence,
            publication.source_rows,
            publication.filtered_rows,
            publication.validation_status,
            publication.quality_ready,
            created_at,
            WORKBOOK_VERSION,
        ]],
        "AnalysisResults": [],
        "AIResponses": [[
            run_id,
            response.language or publication.response_language,
            response.answer,
            response.primary_finding,
            response.calculation,
            response.limitations,
            response.follow_up_question,
            publication.reasoning_summary,
            _json_text(list(publication.assumptions)),
            _json_text(list(publication.tool_steps)),
            created_at,
        ]],
        "QualityEvidence": [],
        "SourceFiles": [],
    }

    for result_order, result in enumerate(publication.results, start=1):
        rows_by_sheet["AnalysisResults"].append([
            run_id,
            result_order,
            result.label,
            result.value,
            result.rank,
            publication.metric,
            publication.dimension,
            result.unit,
            created_at,
        ])

    for item in quality_rows:
        rows_by_sheet["QualityEvidence"].append([
            run_id,
            item.gate_name,
            item.ready,
            item.duplicate_rows,
            item.missing_values,
            item.usable_numeric_ratio,
            item.validation_status,
            _json_text(list(item.warnings)),
            _json_text(dict(item.evidence)),
            created_at,
        ])

    for item in publication.source_files:
        rows_by_sheet["SourceFiles"].append([
            run_id,
            item.filename,
            item.file_type,
            item.file_size_bytes,
            item.page_count,
            item.ocr_executed,
            item.ocr_languages,
            item.extraction_strategy,
            item.extraction_confidence,
            item.source_sha256,
            created_at,
        ])

    _append_rows(workbook, rows_by_sheet)
    _style_workbook(workbook)
    _atomic_save(workbook, path)

    return PowerBIPublishResult(
        run_id=run_id,
        publication_key=publication_key,
        workbook_path=path.resolve(),
        duplicate=False,
        result_rows_written=len(publication.results),
    )


def result_rows_from_records(
    records: Iterable[Mapping[str, Any]],
    *,
    label_field: str = "label",
    value_field: str = "value",
    rank_field: str = "rank",
    unit: str | None = None,
) -> list[PowerBIResultRow]:
    """Convert executor records into typed Power BI result rows."""

    rows: list[PowerBIResultRow] = []
    for index, record in enumerate(records, start=1):
        if label_field not in record or value_field not in record:
            raise KeyError(
                f"record {index} must contain '{label_field}' and '{value_field}'"
            )
        rank_value = record.get(rank_field)
        rows.append(
            PowerBIResultRow(
                label=str(record[label_field]),
                value=record[value_field],
                rank=int(rank_value) if rank_value is not None else None,
                unit=unit,
            )
        )
    return rows
