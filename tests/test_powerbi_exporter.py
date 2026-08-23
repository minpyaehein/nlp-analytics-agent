from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from app.services.powerbi_exporter import (
    SHEET_HEADERS,
    PowerBIPublication,
    PowerBIQualityEvidence,
    PowerBIResponse,
    PowerBIResultRow,
    PowerBISourceFile,
    publish_to_powerbi_workbook,
)


def sample_publication() -> PowerBIPublication:
    return PowerBIPublication(
        question="Show the top 5 products by revenue.",
        planner_source="local_qwen",
        model_name="qwen3:4b",
        intent="ranking",
        metric="revenue",
        dimension="product",
        aggregation="sum",
        sort_direction="descending",
        result_limit=5,
        visualization="bar",
        confidence=0.95,
        source_rows=12,
        filtered_rows=12,
        validation_status="passed",
        quality_ready=True,
        results=(
            PowerBIResultRow("Laptop", 4250, 1),
            PowerBIResultRow("Monitor", 1200, 2),
            PowerBIResultRow("Mouse", 500, 3),
            PowerBIResultRow("Keyboard", 405, 4),
        ),
        ai_response=PowerBIResponse(
            language="en",
            answer="Laptop has the highest validated revenue.",
            primary_finding="Laptop: 4,250",
            calculation="Revenue = quantity * unit_price",
        ),
        quality=(
            PowerBIQualityEvidence(
                gate_name="revenue",
                ready=True,
                duplicate_rows=1,
                missing_values=1,
                usable_numeric_ratio=1.0,
                validation_status="passed",
                warnings=("One duplicate source row was retained.",),
                evidence={"quantity_ratio": 1.0, "unit_price_ratio": 1.0},
            ),
        ),
        source_files=(
            PowerBISourceFile(
                filename="scanned_sales_report_highres.pdf",
                file_type="pdf",
                file_size_bytes=633985,
                page_count=1,
                ocr_executed=True,
                ocr_languages="eng+mya",
                extraction_strategy="coordinate_reconstruction",
                extraction_confidence=1.0,
                source_sha256="abc123",
            ),
        ),
        reasoning_summary="Rank product revenue descending and return the top five.",
        assumptions=("Revenue is quantity multiplied by unit price.",),
        tool_steps=(
            {"tool": "inspect_schema", "reason": "Verify columns."},
            {"tool": "calculate_ranking", "reason": "Calculate deterministically."},
            {"tool": "generate_chart", "reason": "Create bar chart."},
        ),
    )


def test_creates_powerbi_workbook_with_required_sheets(tmp_path: Path) -> None:
    output = tmp_path / "insightflow_powerbi.xlsx"
    result = publish_to_powerbi_workbook(sample_publication(), output)

    assert output.exists()
    assert result.duplicate is False
    assert result.result_rows_written == 4

    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == list(SHEET_HEADERS)
    for name, headers in SHEET_HEADERS.items():
        actual = [workbook[name].cell(1, index).value for index in range(1, len(headers) + 1)]
        assert actual == headers


def test_writes_numeric_results_and_shared_run_id(tmp_path: Path) -> None:
    output = tmp_path / "insightflow_powerbi.xlsx"
    published = publish_to_powerbi_workbook(sample_publication(), output)
    workbook = load_workbook(output, data_only=True)

    result_sheet = workbook["AnalysisResults"]
    assert result_sheet.max_row == 5
    assert result_sheet["A2"].value == published.run_id
    assert result_sheet["C2"].value == "Laptop"
    assert result_sheet["D2"].value == 4250
    assert isinstance(result_sheet["D2"].value, (int, float))

    for sheet_name in ("AIResponses", "QualityEvidence", "SourceFiles"):
        assert workbook[sheet_name]["A2"].value == published.run_id


def test_duplicate_publication_is_idempotent(tmp_path: Path) -> None:
    output = tmp_path / "insightflow_powerbi.xlsx"
    first = publish_to_powerbi_workbook(sample_publication(), output)
    second = publish_to_powerbi_workbook(sample_publication(), output)

    assert first.duplicate is False
    assert second.duplicate is True
    assert second.run_id == first.run_id
    assert second.result_rows_written == 0

    workbook = load_workbook(output)
    assert workbook["AnalysisRuns"].max_row == 2
    assert workbook["AnalysisResults"].max_row == 5


def test_preserves_myanmar_text(tmp_path: Path) -> None:
    output = tmp_path / "insightflow_powerbi.xlsx"
    publication = sample_publication()
    publication = PowerBIPublication(
        **{
            **publication.__dict__,
            "question": "ရန်ကုန်ဒေသ၏ ဝင်ငွေကို ပြပါ",
            "response_language": "my",
            "ai_response": PowerBIResponse(
                language="my",
                answer="ရန်ကုန်ဒေသ၏ အတည်ပြုပြီးသော ဝင်ငွေမှာ ၄,၄၁၅ ဖြစ်သည်။",
            ),
            "publication_key": "myanmar-test",
        }
    )
    publish_to_powerbi_workbook(publication, output)
    workbook = load_workbook(output, data_only=True)

    assert workbook["AnalysisRuns"]["C2"].value == "ရန်ကုန်ဒေသ၏ ဝင်ငွေကို ပြပါ"
    assert "ရန်ကုန်" in workbook["AIResponses"]["C2"].value


def test_quality_json_is_valid(tmp_path: Path) -> None:
    output = tmp_path / "insightflow_powerbi.xlsx"
    publish_to_powerbi_workbook(sample_publication(), output)
    workbook = load_workbook(output, data_only=True)

    warnings = json.loads(workbook["QualityEvidence"]["H2"].value)
    evidence = json.loads(workbook["QualityEvidence"]["I2"].value)
    assert warnings == ["One duplicate source row was retained."]
    assert evidence["quantity_ratio"] == 1.0
